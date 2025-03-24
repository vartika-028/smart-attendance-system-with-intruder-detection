import cv2
import numpy as np
import face_recognition as face_rec
import os
import pyttsx3 as textSpeach
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Initialize text-to-speech engine
engine = textSpeach.init()

# Paths
dataset_path = r"C:\Users\varti\OneDrive\Desktop\project\student_images"
attendance_path = r"C:\Users\varti\OneDrive\Desktop\project\attendance.xlsx"
intruder_log_path = r"C:\Users\varti\Downloads\face_recogniton-learnFromBasics-main\intruders.xlsx"
intruder_images_dir = r"C:\Users\varti\Downloads\face_recogniton-learnFromBasics-main\intruders"

# Ensure directories and Excel files exist
os.makedirs(intruder_images_dir, exist_ok=True)

for path, headers in [
    (attendance_path, ["Name", "Time"]),
    (intruder_log_path, ["Timestamp", "Image_Path"])
]:
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(path)

# Load dataset and encode images
def load_and_encode_dataset(dataset_path):
    known_encodings = []
    known_names = []
    for person_name in os.listdir(dataset_path):
        person_folder = os.path.join(dataset_path, person_name)
        if os.path.isdir(person_folder):  # Ensure it's a directory
            print(f"Processing {person_name}...")
            for filename in os.listdir(person_folder):
                file_path = os.path.join(person_folder, filename)
                if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                    img = cv2.imread(file_path)
                    if img is not None:
                        try:
                            img = cv2.resize(img, (0, 0), fx=0.5, fy=0.5)
                            rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                            encodings = face_rec.face_encodings(rgb_img)
                            if encodings:
                                known_encodings.append(encodings[0])
                                known_names.append(person_name)
                            else:
                                print(f"No face found in {file_path}. Skipping.")
                        except Exception as e:
                            print(f"Error processing {file_path}: {e}")
    return known_encodings, known_names

# Mark attendance in Excel
def mark_attendance(name):
    wb = load_workbook(attendance_path)
    ws = wb.active

    # Check if already marked today
    recorded_names = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]
    if name not in recorded_names:
        now = datetime.now()
        time_str = now.strftime('%H:%M:%S')
        ws.append([name, time_str])
        wb.save(attendance_path)
        print(f"Attendance marked for {name} at {time_str}")

        # Announce attendance
        engine.say(f"Welcome to class, {name}")
        engine.runAndWait()

# Log intruder in Excel and save image
def log_intruder(frame):
    wb = load_workbook(intruder_log_path)
    ws = wb.active

    now = datetime.now()
    timestamp = now.strftime('%Y-%m-%d %H:%M:%S')
    img_name = f"intruder_{now.strftime('%Y%m%d_%H%M%S')}.jpg"
    img_path = os.path.join(intruder_images_dir, img_name)

    # Save intruder image
    cv2.imwrite(img_path, frame)

    # Log details in Excel
    ws.append([timestamp, img_path])
    wb.save(intruder_log_path)

    # Announce intruder alert
    engine.say("Intruder detected! Snapshot taken.")
    engine.runAndWait()

# Load dataset
print("Loading dataset...")
known_encodings, known_names = load_and_encode_dataset(dataset_path)
print(f"Loaded {len(known_encodings)} faces.")

# Initialize webcam
vid = cv2.VideoCapture(0)

try:
    while True:
        success, frame = vid.read()
        if not success:
            print("Failed to capture video frame. Exiting...")
            break

        # Resize frame for faster processing
        smaller_frame = cv2.resize(frame, (0, 0), None, 0.25, 0.25)
        rgb_smaller_frame = cv2.cvtColor(smaller_frame, cv2.COLOR_BGR2RGB)

        # Detect faces and encodings in the current frame
        faces_in_frame = face_rec.face_locations(rgb_smaller_frame)
        encodings_in_frame = face_rec.face_encodings(rgb_smaller_frame, faces_in_frame)

        for encode_face, face_loc in zip(encodings_in_frame, faces_in_frame):
            # Match the detected face with known encodings
            matches = face_rec.compare_faces(known_encodings, encode_face)
            face_distance = face_rec.face_distance(known_encodings, encode_face)
            match_index = np.argmin(face_distance)

            if matches[match_index]:
                name = known_names[match_index].upper()
                y1, x2, y2, x1 = face_loc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4

                # Draw a rectangle and name label
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 3)
                cv2.rectangle(frame, (x1, y2 - 25), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(frame, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

                # Mark attendance
                mark_attendance(name)
            else:
                # Intruder detected
                y1, x2, y2, x1 = face_loc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4

                # Draw a rectangle and label as "UNKNOWN"
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 3)
                cv2.rectangle(frame, (x1, y2 - 25), (x2, y2), (0, 0, 255), cv2.FILLED)
                cv2.putText(frame, "UNKNOWN", (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

                # Log intruder details and save snapshot
                intruder_frame = frame.copy()
                log_intruder(intruder_frame)

        # Display the video frame
        cv2.imshow("Face Recognition", frame)

        # Exit condition: Press 'q' to quit
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

finally:
    # Release resources
    vid.release()
    cv2.destroyAllWindows()
    print("Resources released.")
